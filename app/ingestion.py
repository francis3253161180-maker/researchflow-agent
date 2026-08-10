from __future__ import annotations

import re
import statistics
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

from docx import Document
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
import pymupdf
from openpyxl import load_workbook
from pypdf import PdfReader


SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".md", ".txt"}


@dataclass(frozen=True)
class TextBlock:
    content: str
    page: int | None = None
    section: str | None = None


@dataclass(frozen=True)
class ParsedDocument:
    title: str
    source: str
    filename: str
    media_type: str
    content: str
    blocks: list[TextBlock]


def _normalize(text: str) -> str:
    # PDF extractors occasionally emit lone UTF-16 surrogates. SQLite's UTF-8
    # encoder rejects them, so normalize them at the ingestion boundary instead
    # of letting one malformed glyph invalidate an otherwise usable document.
    safe_text = text.encode("utf-8", errors="replace").decode("utf-8")
    return re.sub(r"\n{3,}", "\n\n", safe_text.replace("\r\n", "\n")).strip()


def _section_path(stack: list[tuple[int, str]]) -> str | None:
    return " › ".join(item[1] for item in stack) or None


def _update_heading_stack(stack: list[tuple[int, str]], level: int, heading: str) -> None:
    """Replace the active branch while preserving parent heading context."""
    compact = _normalize(heading).replace("\n", " ")
    if not compact:
        return
    level = max(1, min(level, 9))
    if stack and stack[-1][1] == compact and stack[-1][0] == level:
        return
    while stack and stack[-1][0] >= level:
        stack.pop()
    stack.append((level, compact))


def _markdown_blocks(text: str) -> list[TextBlock]:
    blocks: list[TextBlock] = []
    heading_stack: list[tuple[int, str]] = []
    buffer: list[str] = []

    def flush() -> None:
        nonlocal buffer
        if buffer:
            blocks.append(TextBlock(_normalize("\n".join(buffer)), section=_section_path(heading_stack)))
            buffer = []

    for line in text.splitlines():
        heading = re.match(r"^\s*(#{1,6})\s+(.+?)\s*#*\s*$", line)
        if heading:
            flush()
            _update_heading_stack(heading_stack, len(heading.group(1)), heading.group(2))
        else:
            buffer.append(line)
    flush()
    return [block for block in blocks if block.content]


def _docx_heading_level(paragraph: Paragraph) -> int | None:
    """Read built-in Word heading levels without treating arbitrary text as a heading."""
    style = paragraph.style
    seen: set[str] = set()
    while style is not None:
        style_id = str(getattr(style, "style_id", ""))
        name = str(getattr(style, "name", ""))
        signature = f"{style_id}|{name}"
        if signature in seen:
            break
        seen.add(signature)
        match = re.search(r"(?:heading|标题)\s*([1-9]\d*)$", f"{name} {style_id}", re.IGNORECASE)
        if match:
            return int(match.group(1))
        style = getattr(style, "base_style", None)
    return None


def _iter_docx_blocks(document):
    """Yield paragraphs and tables in their actual WordprocessingML order."""
    for child in document.element.body.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, document)
        elif isinstance(child, CT_Tbl):
            yield Table(child, document)


def _table_text(table: Table) -> str:
    rows: list[str] = []
    for row in table.rows:
        cells = [cell.text.strip() for cell in row.cells if cell.text.strip()]
        if cells:
            rows.append(" | ".join(cells))
    return _normalize("\n".join(rows))


def _docx_blocks(payload: bytes) -> list[TextBlock]:
    document = Document(BytesIO(payload))
    blocks: list[TextBlock] = []
    heading_stack: list[tuple[int, str]] = []
    buffer: list[str] = []

    def flush() -> None:
        nonlocal buffer
        content = _normalize("\n\n".join(buffer))
        if content:
            blocks.append(TextBlock(content, section=_section_path(heading_stack)))
        buffer = []

    for item in _iter_docx_blocks(document):
        if isinstance(item, Paragraph):
            text = _normalize(item.text)
            if not text:
                continue
            heading_level = _docx_heading_level(item)
            if heading_level is not None:
                flush()
                _update_heading_stack(heading_stack, heading_level, text)
            else:
                buffer.append(text)
        else:
            flush()
            content = _table_text(item)
            if content:
                blocks.append(TextBlock(content, section=_section_path(heading_stack)))
    flush()
    return blocks


@dataclass(frozen=True)
class _PdfLine:
    text: str
    size: float
    bold: bool
    y0: float
    y1: float


def _pdf_lines(page) -> list[_PdfLine]:
    lines: list[_PdfLine] = []
    for block in page.get_text("dict", sort=True).get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            spans = line.get("spans", [])
            text = _normalize("".join(str(span.get("text", "")) for span in spans))
            if not text:
                continue
            sizes = [float(span.get("size", 0.0)) for span in spans if span.get("text", "").strip()]
            if not sizes:
                continue
            fonts = " ".join(str(span.get("font", "")) for span in spans).lower()
            bbox = line.get("bbox", (0.0, 0.0, 0.0, 0.0))
            lines.append(_PdfLine(text, max(sizes), "bold" in fonts or "black" in fonts, float(bbox[1]), float(bbox[3])))
    return lines


def _numbered_heading_level(text: str) -> int | None:
    match = re.match(r"^\s*(\d+(?:\.\d+){0,5}|[IVXLCDM]+|[A-Z])(?:[.)])?\s+\S", text)
    if not match:
        return None
    label = match.group(1)
    return len(label.split(".")) if label[0].isdigit() else 1


def _pdf_heading_level(line: _PdfLine, body_size: float, repeated_margin_text: set[str]) -> int | None:
    text = line.text
    normalized = " ".join(text.lower().split())
    if normalized in repeated_margin_text or len(text) > 180:
        return None
    numbered_level = _numbered_heading_level(text)
    font_ratio = line.size / max(body_size, 1.0)
    # Numbering is strong structural evidence; unnumbered headings require a
    # clearer layout signal to avoid treating ordinary short prose as a title.
    if numbered_level is not None and (line.bold or font_ratio >= 1.05):
        return numbered_level
    if font_ratio >= 1.55 and len(text) <= 120:
        return 1
    if line.bold and font_ratio >= 1.25 and len(text) <= 100:
        return 2
    return None


def _pdf_blocks(payload: bytes) -> list[TextBlock]:
    document = pymupdf.open(stream=payload, filetype="pdf")
    try:
        pages = [_pdf_lines(page) for page in document]
        body_sizes = [line.size for page_lines in pages for line in page_lines if len(line.text) >= 30]
        body_size = statistics.median(body_sizes) if body_sizes else 10.0
        margin_occurrences: dict[str, set[int]] = {}
        for page_number, (page, lines) in enumerate(zip(document, pages), start=1):
            for line in lines:
                if line.y0 <= 72 or line.y1 >= page.rect.height - 72:
                    margin_occurrences.setdefault(" ".join(line.text.lower().split()), set()).add(page_number)
        repeated_margin_text = {text for text, seen_pages in margin_occurrences.items() if len(seen_pages) >= 2}

        blocks: list[TextBlock] = []
        heading_stack: list[tuple[int, str]] = []
        for page_number, page_lines in enumerate(pages, start=1):
            buffer: list[str] = []

            def flush() -> None:
                nonlocal buffer
                content = _normalize("\n".join(buffer))
                if content:
                    blocks.append(TextBlock(content, page=page_number, section=_section_path(heading_stack)))
                buffer = []

            for line in page_lines:
                level = _pdf_heading_level(line, body_size, repeated_margin_text)
                if level is not None:
                    flush()
                    _update_heading_stack(heading_stack, level, line.text)
                else:
                    buffer.append(line.text)
            flush()
        return blocks
    finally:
        document.close()


def _spreadsheet_blocks(payload: bytes) -> list[TextBlock]:
    """Serialize an XLSX workbook as evidence-bearing table rows.

    Each block records its worksheet and inclusive row range. Formulas are
    preserved as formulas rather than executed; macros are never run. This
    makes spreadsheet evidence inspectable without claiming spreadsheet-agent
    capabilities that the service does not provide.
    """
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=False)
    blocks: list[TextBlock] = []
    for sheet in workbook.worksheets:
        rows: list[str] = []
        headers: list[str] = []
        start_row: int | None = None
        end_row: int | None = None

        def flush() -> None:
            nonlocal rows, start_row, end_row
            if rows and start_row is not None and end_row is not None:
                section = f"工作表：{sheet.title}｜行 {start_row}-{end_row}"
                blocks.append(TextBlock(_normalize("\n".join(rows)), section=section))
            rows = []
            start_row = None
            end_row = None

        for row_number, cells in enumerate(sheet.iter_rows(), start=1):
            values = ["" if cell.value is None else str(cell.value).strip() for cell in cells]
            if not any(values):
                headers = []
                continue
            nonempty = [value for value in values if value]
            text_like = [value for value in nonempty if not re.fullmatch(r"[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?(?:±\d+(?:\.\d*)?)?", value)]
            # Workbooks often contain a few parameter-value records before a
            # real table. A short row is ambiguous, so only adopt a header when
            # it has several text labels; otherwise retain neutral column IDs.
            if len(nonempty) >= 4 and len(text_like) >= 4:
                headers = [value or f"列{index}" for index, value in enumerate(values, start=1)]
                rows.append(f"行 {row_number}｜表头：" + "；".join(f"列{index + 1}：{value}" for index, value in enumerate(values) if value))
                start_row = row_number if start_row is None else start_row
                end_row = row_number
                continue
            fields = [
                f"{headers[index] if index < len(headers) else f'列{index + 1}'}：{value}"
                for index, value in enumerate(values)
                if value
            ]
            if not fields:
                continue
            rows.append(f"行 {row_number}｜" + "；".join(fields))
            start_row = row_number if start_row is None else start_row
            end_row = row_number
            # Keep blocks compact enough that citations can point to a useful
            # row range before generic text chunking applies its own split.
            if len(rows) >= 40:
                flush()
        flush()
    return blocks


def parse_upload(filename: str, payload: bytes, source: str = "upload") -> ParsedDocument:
    suffix = Path(filename).suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise ValueError("only PDF, DOCX, XLSX, Markdown, and TXT files are supported")
    if not payload:
        raise ValueError("uploaded file is empty")

    title = Path(filename).stem[:200] or "untitled"
    if suffix == ".pdf":
        try:
            blocks = _pdf_blocks(payload)
        except Exception:
            # Preserve the existing text-layer fallback for malformed PDFs or
            # rare parser incompatibilities. It retains page numbers but does
            # not claim a section when layout metadata was unavailable.
            reader = PdfReader(BytesIO(payload))
            blocks = [TextBlock(_normalize(page.extract_text() or ""), page=index) for index, page in enumerate(reader.pages, start=1)]
        blocks = [block for block in blocks if block.content]
        media_type = "application/pdf"
    elif suffix == ".docx":
        blocks = _docx_blocks(payload)
        media_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    elif suffix == ".xlsx":
        try:
            blocks = _spreadsheet_blocks(payload)
        except Exception as exc:
            raise ValueError("could not read XLSX workbook; encrypted or malformed files are not supported") from exc
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        text = _normalize(payload.decode("utf-8", errors="replace"))
        blocks = _markdown_blocks(text) if suffix == ".md" else [TextBlock(text)]
        media_type = "text/markdown" if suffix == ".md" else "text/plain"

    content = _normalize("\n\n".join(block.content for block in blocks))
    if len(content) < 20:
        raise ValueError("no extractable text was found in the uploaded file")
    return ParsedDocument(title, source, filename, media_type, content, blocks)
